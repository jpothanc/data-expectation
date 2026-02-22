"""
Excel report generation service for validation failure data.

Architecture
------------
Module-level pure functions handle all openpyxl styling and sheet construction.
``ReportService`` is the single public class: it owns the data-access layer
(analytics + instruments) and delegates sheet-building to the helpers below.

Color scheme: clean light / white theme.
- White and very-light-blue row backgrounds
- Navy blue column headers
- Amber/orange section banners for failure callouts
- Dark text throughout for readability
"""

import json
import logging

import pandas as pd
from io import BytesIO

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Instrument columns to include in every failed-check instrument table.
# The failing column is appended automatically if it is not already present.
INSTRUMENT_DETAIL_COLS = ['MasterId', 'RIC', 'Sedol', 'Exchange']


# ── Style palette ──────────────────────────────────────────────────────────

def _fill(h: str) -> PatternFill:
    return PatternFill('solid', fgColor=h)


def _font(h: str, bold: bool = False, sz: int = 10) -> Font:
    return Font(color=h, bold=bold, size=sz)


def _border() -> Border:
    t = Side(style='thin', color='D1D5DB')
    return Border(left=t, right=t, top=t, bottom=t)


# ── Light / white color palette ────────────────────────────────────────────

# Backgrounds
FILL_TITLE   = _fill('1E3A5F')   # navy — report title banner
FILL_HDR     = _fill('2563A8')   # medium blue — column headers
FILL_ROW     = _fill('FFFFFF')   # white — body rows
FILL_ROW_ALT = _fill('EFF6FF')   # very light blue — alternating rows
FILL_SECTION = _fill('FEF3C7')   # light amber — failed-check section banner
FILL_WARN    = _fill('FFFBEB')   # pale amber — failed-check stats / instrument rows

# Fonts
FONT_TITLE   = _font('FFFFFF', bold=True,  sz=12)  # white on navy title
FONT_HDR     = _font('FFFFFF', bold=True,  sz=10)  # white on blue header
FONT_BODY    = _font('1E293B', sz=10)              # dark slate — body text
FONT_MUTED   = _font('64748B', sz=10)              # mid slate — secondary info
FONT_WARN    = _font('92400E', sz=10)              # dark amber — failure values
FONT_WARN_B  = _font('92400E', bold=True, sz=10)   # dark amber bold — banners

AL  = Alignment(horizontal='left',   vertical='center')
AC  = Alignment(horizontal='center', vertical='center', wrap_text=True)
BDR = _border()


# ── Low-level cell helpers ─────────────────────────────────────────────────

def _cell(ws, row: int, col: int, value, bg=None, font=None, align=None):
    c = ws.cell(row=row, column=col, value=value)
    if bg:   c.fill      = bg
    if font: c.font      = font
    c.alignment = align or AL
    c.border    = BDR
    return c


def _header_row(ws, row_idx: int, headers: list, bg=None):
    for ci, text in enumerate(headers, 1):
        _cell(ws, row_idx, ci, text, bg or FILL_HDR, FONT_HDR, AC)


def _banner(ws, row: int, ncols: int, text: str,
            bg, font, align=AL, height: int = 20):
    """Write a merged full-width banner row."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = bg; c.font = font; c.alignment = align; c.border = BDR
    ws.row_dimensions[row].height = height
    return c


def _autofit(ws, min_w: int = 10, max_w: int = 55):
    for col_cells in ws.columns:
        w = max((len(str(cc.value or '')) for cc in col_cells), default=0)
        ws.column_dimensions[
            get_column_letter(col_cells[0].column)
        ].width = min(max(w + 2, min_w), max_w)


# ── Data helpers ───────────────────────────────────────────────────────────

def _parse_result_details(raw) -> tuple[list[str], bool]:
    """
    Extract unexpected-value list and missing-flag from a GE ResultDetails blob.

    Returns
    -------
    (values, include_missing)
        values          – list of string values that failed the check
        include_missing – True when missing/null rows should also be fetched
    """
    if not raw:
        return [], False
    try:
        data = json.loads(str(raw))
        counts = data.get('partial_unexpected_counts') or []
        if counts:
            return [str(c['value']) for c in counts if c.get('value') is not None], False
        lst = data.get('partial_unexpected_list') or []
        if lst:
            return [str(v) for v in lst if v is not None], False
        return [], bool((data.get('missing_count') or 0) > 0)
    except Exception:
        return [], False


def _filter_instruments(df, column: str,
                        values: list[str],
                        include_missing: bool) -> list[dict]:
    """
    Filter *df* to rows where *column* is in *values* or is null/empty.

    Returns a list of dicts limited to INSTRUMENT_DETAIL_COLS + the failing
    column (appended only when it is not already in the standard set).
    """
    if df is None or df.empty or column not in df.columns:
        return []

    mask = pd.Series(False, index=df.index)
    if values:
        mask |= df[column].astype(str).isin([str(v) for v in values])
    if include_missing:
        mask |= df[column].isna() | (df[column].astype(str).str.strip() == '')

    matched = df[mask]
    if matched.empty:
        return []

    export_cols = [c for c in INSTRUMENT_DETAIL_COLS if c in matched.columns]
    if column not in export_cols and column in matched.columns:
        export_cols.append(column)

    subset = matched[export_cols].astype(object).where(
        pd.notnull(matched[export_cols]), other=None
    )
    return subset.to_dict(orient='records')


# ── Sheet builders (pure functions, take no I/O) ───────────────────────────

def _build_summary_sheet(ws, runs: list, region: str, date: str):
    """Populate the Summary sheet with one row per failed exchange."""
    ws.title = 'Summary'
    ws.sheet_view.showGridLines = False

    _banner(ws, 1, 7,
            f'Validation Failure Report  |  {region.upper()}  |  {date}',
            FILL_TITLE, FONT_TITLE, AC, height=28)
    _banner(ws, 2, 7,
            f'Failed exchanges: {len(runs)}',
            FILL_TITLE, _font('BFD7ED', sz=10), AC, height=16)
    for ci in range(1, 8):
        ws.cell(row=3, column=ci).fill = FILL_TITLE

    _header_row(ws, 4, [
        'Exchange', 'Product Type', 'Run Timestamp',
        'Failed Checks', 'Pass Rate %', 'Duration (ms)', 'Rules Applied',
    ])
    ws.freeze_panes = 'A5'

    for dr, run in enumerate(runs, start=5):
        tot  = run.get('TotalExpectations', 0) or 0
        fail = run.get('FailedExpectations', 0) or 0
        pct  = round((tot - fail) / tot * 100, 1) if tot > 0 else 0.0
        bg   = FILL_WARN if fail > 0 else FILL_ROW_ALT
        _cell(ws, dr, 1, run.get('Exchange', ''),                bg, FONT_BODY)
        _cell(ws, dr, 2, run.get('ProductType', ''),             bg, FONT_BODY)
        _cell(ws, dr, 3, str(run.get('RunTimestamp', '')),       bg, FONT_MUTED)
        _cell(ws, dr, 4, fail,                                   bg, FONT_WARN_B, AC)
        _cell(ws, dr, 5, f'{pct}%',                              bg, FONT_WARN,   AC)
        _cell(ws, dr, 6, run.get('ExecutionDurationMs', 0) or 0, bg, FONT_MUTED,  AC)
        _cell(ws, dr, 7, len(run.get('rules_applied', [])),      bg, FONT_MUTED,  AC)

    _autofit(ws)


def _build_exchange_sheet(ws, run: dict, get_df_fn):
    """
    Populate one exchange tab with a section per failed expectation,
    each section containing the matching instrument details.

    Parameters
    ----------
    ws         – blank worksheet to populate
    run        – single run dict from the analytics service
    get_df_fn  – callable(exchange, product_type) -> pd.DataFrame | None
    """
    ws.sheet_view.showGridLines = False

    exchange = run.get('Exchange', 'Unknown')
    product  = run.get('ProductType', 'unknown')
    tot      = run.get('TotalExpectations', 0) or 0
    fail     = run.get('FailedExpectations', 0) or 0
    pct      = round((tot - fail) / tot * 100, 1) if tot > 0 else 0.0
    ts       = run.get('RunTimestamp', '')
    dur      = run.get('ExecutionDurationMs', 0) or 0

    # Normalise product type for instrument service (e.g. "stocks" → "stock")
    norm_pt = product.lower().rstrip('s') if product.lower().endswith('s') else product.lower()

    # Tab header
    _banner(ws, 1, 7,
            f'{exchange}  |  {product}  |  Failed Rule Details',
            FILL_TITLE, FONT_TITLE, height=22)
    _banner(ws, 2, 7,
            f'Run: {ts}  |  Failed: {fail} / {tot}  |  Pass rate: {pct}%  |  Duration: {dur} ms',
            FILL_SECTION, FONT_WARN, height=16)

    failed_exps = [e for e in run.get('expectation_results', [])
                   if not e.get('Success', True)]

    cur = 4  # current row pointer

    if not failed_exps:
        _banner(ws, cur, 7, '  No failed expectation details recorded.',
                FILL_ROW_ALT, FONT_MUTED, height=16)
        _autofit(ws)
        return

    # Load exchange DataFrame once — shared across all failed checks on this tab
    df = get_df_fn(exchange, norm_pt)

    for idx, exp in enumerate(failed_exps):
        col_name  = exp.get('ColumnName', '')
        exp_type  = exp.get('ExpectationType', '')
        unexp_cnt = exp.get('UnexpectedCount', 0) or 0
        unexp_pct = round(exp.get('UnexpectedPercent', 0) or 0, 2)
        miss_cnt  = exp.get('MissingCount', 0) or 0

        values, inc_missing = _parse_result_details(exp.get('ResultDetails', ''))
        instruments         = _filter_instruments(df, col_name, values, inc_missing)

        # ── Section banner ─────────────────────────────────────────────────
        _banner(ws, cur, 7,
                f'  FAILED CHECK {idx + 1} / {len(failed_exps)}'
                f'   |   Column: {col_name}   |   {exp_type}',
                FILL_SECTION, FONT_WARN_B, height=18)
        cur += 1

        # ── Stats strip ────────────────────────────────────────────────────
        _banner(ws, cur, 7,
                f'  Unexpected: {unexp_cnt} ({unexp_pct}%)'
                f'   |   Missing: {miss_cnt}'
                f'   |   Instruments found: {len(instruments)}',
                FILL_WARN, FONT_WARN, height=15)
        cur += 1

        # ── Instrument table ───────────────────────────────────────────────
        if instruments:
            inst_cols = list(instruments[0].keys())
            _header_row(ws, cur, inst_cols)
            ws.freeze_panes = f'A{cur + 1}'
            cur += 1

            for ii, inst in enumerate(instruments):
                bg = FILL_ROW if ii % 2 == 0 else FILL_ROW_ALT
                for ci, key in enumerate(inst_cols, 1):
                    is_fail_col = (key == col_name)
                    _cell(ws, cur, ci,
                          inst.get(key, ''),
                          bg,
                          FONT_WARN if is_fail_col else FONT_BODY,
                          AC if key in ('Exchange', 'Sedol') else AL)
                cur += 1
        else:
            _banner(ws, cur, 7,
                    '  No matching instruments found '
                    '(data may have changed since validation ran).',
                    FILL_ROW_ALT, FONT_MUTED, height=15)
            cur += 1

        cur += 1  # blank gap between sections

    _autofit(ws)


# ── Public service class ───────────────────────────────────────────────────

class ReportService:
    """
    Orchestrates data retrieval and Excel workbook generation for validation reports.

    Responsibilities
    ----------------
    - Fetch validation run data via ``ValidationAnalyticsService``
    - Lazily initialise and cache ``InstrumentService`` instances per product type
    - Cache exchange DataFrames for the duration of one report generation
    - Delegate all sheet-building to the module-level pure functions above

    The controller should only call ``generate_excel_report`` and handle the
    HTTP response — no workbook or pandas logic belongs there.
    """

    def __init__(self):
        from services.validation_analytics_service import ValidationAnalyticsService
        from services.loader_factory import LoaderFactory
        self._analytics       = ValidationAnalyticsService()
        self._loader_factory  = LoaderFactory()
        self._instrument_svcs = {}   # product_type -> InstrumentService (long-lived)

    # ── Private helpers ────────────────────────────────────────────────────

    def _instrument_svc(self, product_type: str):
        """Return (and cache) an InstrumentService for *product_type*."""
        if product_type not in self._instrument_svcs:
            from services.instrument_service import InstrumentService
            loader = self._loader_factory.create_loader()
            exmap  = self._loader_factory.get_exchange_map(product_type=product_type)
            self._instrument_svcs[product_type] = InstrumentService(
                loader, exchange_map=exmap, product_type=product_type
            )
        return self._instrument_svcs[product_type]

    def _make_df_loader(self) -> dict:
        """
        Return a per-request (exchange, product_type) → DataFrame cache plus the
        loader function that populates it.  Each call to ``generate_excel_report``
        gets a fresh cache so stale data never carries over between requests.
        """
        cache = {}

        def get_df(exchange: str, product_type: str):
            key = (exchange.upper(), product_type)
            if key not in cache:
                try:
                    cache[key] = (
                        self._instrument_svc(product_type)
                            .load_exchange_data(exchange.upper())
                    )
                except Exception as exc:
                    logger.warning(
                        'Cannot load instruments for %s / %s: %s',
                        exchange, product_type, exc
                    )
                    cache[key] = None
            return cache[key]

        return get_df

    # ── Public API ─────────────────────────────────────────────────────────

    def generate_excel_report(self, region: str, date: str,
                               days: int = 90) -> BytesIO:
        """
        Build a consolidated Excel workbook for all failed exchanges in
        *region* on *date*.

        Sheet layout
        ------------
        - Sheet 1 "Summary"   : headline metrics for every failed exchange
        - One sheet per exchange: one section per failed expectation containing
          the failing instruments (MasterId, RIC, Sedol, Exchange)

        Returns
        -------
        BytesIO
            Buffer positioned at offset 0, ready for ``send_file``.
        """
        result = self._analytics.get_validation_results_by_region_date(
            region=region, date=date, days=days, limit=500
        )
        runs = result.get('runs', [])

        get_df = self._make_df_loader()

        wb = openpyxl.Workbook()
        _build_summary_sheet(wb.active, runs, region, date)

        used_names = {'Summary'}
        for run in runs:
            exchange   = run.get('Exchange', 'Unknown')
            product    = run.get('ProductType', 'unknown')
            raw_name   = f'{exchange}_{product}'
            sheet_name = raw_name[:31]
            sfx = 2
            while sheet_name in used_names:
                sheet_name = f'{raw_name[:28]}_{sfx}'
                sfx += 1
            used_names.add(sheet_name)

            ws = wb.create_sheet(title=sheet_name)
            _build_exchange_sheet(ws, run, get_df)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
